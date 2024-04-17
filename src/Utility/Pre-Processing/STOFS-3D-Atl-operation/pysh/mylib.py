#!/usr/bin/env python3
#!/usr/bin/evn python3
from pylib import *

#-------misc-------------------------------------------------------------------
def rtext(x,y,note,xm=None,ym=None,ax=None,**args):
    '''
    add annoation to the current axes at relative location to x-axis and y-axis
       x: relative location in x-axis, tpycially 0-1, but can be negative or larger than 1 
       y: relative location in y-axis, typically 0-1, but can be negative or larger than 1
       note: annotation string
       xm: range of x-axis; gca().xlim() is used if not given
       ym: range of y-axis; gca().ylim() is used if not given
       ax: axes; gca() is used if not given
       **args: all other parameters used in matploblib.plot.text are also applicable

    E.q., rtext(0.1,0.9,"(a)",size=10,weight='bold') -> place label at top left corner

    note: suggest not to rearrange the x/ylim after using this function, or place this after
    '''
    if ax is None: ax=gca()
    sca(ax)
    if xm is None: xm=xlim()
    if ym is None: ym=ylim()
    text(xm[0]+x*diff(xm),ym[0]+y*diff(ym)*y,note,**args)
    
def read_excel(fname,sht=0,fmt=0):
    '''
    use pd.read_excel to read Excel file
      fname: name of Excel file
      sht:   name of sheet_name, or number of sheet (default is 0)
      fmt=0: not treat 1st line as header, return data only; fmt=1: treat 1st line as header, return header and data
    '''
    if fmt==0:
       fd=pd.read_excel(fname,sheet_name=sht,header=None)
    else:
       fd=pd.read_excel(fname,sheet_name=sht)
    header=array(fd.columns); fdata=fd.values

    if fmt==0: return fdata
    if fmt==1: return header, fdata

def write_excel(data,fname,sht='sheet_1',indy=0,indx=0,fmt=0,align='row',old_font=0,
                color=None,fontsize=None,fontweight=None,**args):
    '''
    use xlsxwriter to write Excel file
       data: can be a single data, 1D array or 2D array
       fname: name of Excel file
       sht:  name of sheet_name
       indy: starting Row index of cell for data to be written
       indx: starting Column index of cell for data to be written
       align='row': write 1D data as a row; align='column': write 1D data as a column
       old_font=1: for existing sheet,keep old font styles; old_font=0: discard
       color,fontsize,fontweight: options for specify cell font (see openpyxl.styles.Font)

       fmt=0: append data to existing file, or create file if not existing
       fmt=1: replace mode of excel file
       fmt=2: only replace sheet, but keep other sheets of excel file
    '''
    import openpyxl

    #open fname
    if not fname.endswith('.xlsx'): fname=fname+'.xlsx'
    if os.path.exists(fname) and (fmt==0 or fmt==2):
        fid=pd.ExcelWriter(fname,mode='a',engine='openpyxl',if_sheet_exists='replace')
    else:
        fid=pd.ExcelWriter(fname,mode='w+',engine='openpyxl')

    #reorganize data to a 2D array
    if type(data)==str or (not hasattr(data,'__len__')): data=array([[data,],]) #for single data
    data=array(data) #for list
    if data.ndim==1: data=data[None,:] if (align=='row') else data[:,None]  #for 1D array

    #get dimension info
    ds=data.shape; ny, nx=ds[0]+indy, ds[1]+indx

    #create matrix of fields
    if (sht in list(fid.sheets)) and fmt==0:
        data0=read_excel(fname,sht).astype('O'); ny0,nx0=data0.shape
        if old_font==1:
           sid=fid.sheets[sht]; fonts=ones([ny0,nx0]).astype('O')
           for k in arange(ny0):
               for i in arange(nx0): fonts[k,i]=sid.cell(k+1,i+1).font.copy()
        if ny0<ny: data0=r_[data0,tile('',[(ny-ny0),nx0])]; ny0=ny
        if nx0<nx: data0=c_[data0,tile('',[ny0,(nx-nx0)])]; nx0=nx
    else:
        data0=tile('',[ny,nx]).astype('O')

    #write data
    for k, datai in enumerate(data):
        for i, dataii in enumerate(datai): data0[indy+k,indx+i]=dataii

    df=pd.DataFrame([list(i) for i in data0])
    df.to_excel(fid,sheet_name=sht,header=False,index=False)

    #specify cell font
    if old_font==1 and (sht in list(fid.sheets)) and fmt==0:
       sid=fid.sheets[sht]
       for k in arange(fonts.shape[0]):
           for i in arange(fonts.shape[1]): sid.cell(k+1,i+1).font=fonts[k,i]
    if (color,fontsize,fontweight)!=(None,None,None) or len(args)!=0:
       #from openpyxl.styles import Color, PatternFill, Font, Border
       if color is not None: color=mpl.colors.to_hex(color)[1:]
       cf=openpyxl.styles.Font(color=color,size=fontsize,bold=(fontweight=='bold'),**args)
       sid=fid.sheets[sht]
       for k, datai in enumerate(data):
           for i, dataii in enumerate(datai): sid.cell(indy+k+1,indx+i+1).font=cf

    fid.close()

def read_yaml(fname):
    '''
    read yaml file and return key-value dict
    
    '''
    lines=open(fname,'r').readlines()
    param={}
    for line in lines:
        sline=line.strip().split(':')
        if len(sline)<2: continue
        key=sline[0]; value=sline[1].strip()
        if value=='': continue
        value=value.split()[0]
        param[key]=value

    return param

def get_hpc_command(code,bdir,jname='mpi4py',qnode='x5672',nnode=1,ppn=1,wtime='01:00:00',
                    scrout='screen.out',fmt=0,ename='param',qname='flex',account='gg0028'):
    '''
    get command for batch jobs on sciclone/ches/viz3
       code: job script
       bdir: current working directory
       jname: job name
       qname: partition name (needed on some cluster/project)
       qnode: hpc node name
       nnode,ppn,wtime: request node number, core per node, and walltime
       fmt=0: command for submitting batch jobs; fmt=1: command for run parallel jobs
    '''

    nproc=nnode*ppn
    if fmt==0:
       os.environ[ename]='{} {}'.format(bdir,code)
       #for submit jobs
       if qnode in ['femto','cyclops']:
          #scmd='sbatch --export=ALL --constraint=femto --exclusive -J {} -N {} -n {} -t {} {}'.format(jname,nnode,nproc,wtime,code)
          scmd='sbatch --export=ALL -J {} -N {} --ntasks-per-node {} -t {} {}'.format(jname,nnode,ppn,wtime,code)
       elif qnode in ['frontera',]:
          #scmd='sbatch --export=ALL,{}="{} {}" -J {} -p {} -N {} -n {} -t {} {}'.format(ename,bdir,code,jname,qname,nnode,nproc,wtime,code)
          scmd='sbatch --export=ALL -J {} -p {} -N {} --ntasks-per-node {} -t {} {}'.format(jname,qname,nnode,ppn,wtime,code)
       elif qnode in ['mistral',]:
          #scmd='sbatch --export=ALL -J {} --partition=compute2 --account={} -N {} --ntasks-per-node {} -t {} {}'.format(jname,account,nnode,ppn,wtime,code)
          scmd='sbatch --export=ALL -J {} -p {} --account={} -N {} --ntasks-per-node {} -t {} {}'.format(jname,qname,account,nnode,ppn,wtime,code)
       elif qnode in ['stampede2',]:
          scmd='sbatch "--export=ALL" -J {} -p {} -A {} -N {} -n {} -t {} {}'.format(jname,qname,account,nnode,nproc,wtime,code)
       elif qnode in ['x5672','vortex','vortexa','c18x','potomac','james','bora']:
          scmd='qsub {} -v {}="{} {}", -N {} -j oe -l nodes={}:{}:ppn={} -l walltime={}'.format(code,ename,bdir,code,jname,nnode,qnode,ppn,wtime)
          if qnode=='james': scmd='qsub {} -V -v {}="{} {}", -N {} -j oe -l nodes={}:{}:ppn={} -l walltime={}'.format(code,ename,bdir,code,jname,nnode,qnode,ppn,wtime)
       else:
          sys.exit('unknow qnode: {},tag=1'.format(qnode))
    elif fmt==1:
       #for run parallel jobs
       if qnode in ['femto','cyclops']:
          scmd="srun --export=ALL,job_on_node=1,bdir={} ./{} >& {}".format(bdir,code,scrout)
       elif qnode in ['frontera',]:
          scmd="mpirun --env job_on_node 1 --env bdir='{}' -np {} ./{} >& {}".format(bdir,nproc,code,scrout)
          if ename=='run_schism': scmd="ibrun ./{} >& {}".format(code,scrout)
       elif qnode in ['stampede2',]:
          scmd="mpiexec -envall -genv job_on_node 1 -genv bdir '{}' -n {} ./{} >& {}".format(bdir,nproc,code,scrout)
          if ename=='run_schism': scmd="ibrun ./{} >& {}".format(code,scrout)
       elif qnode in ['mistral',]:
          scmd="srun --export=ALL,job_on_node=1,bdir={} -l --propagate=STACK,CORE -l --cpu_bind=verbose,cores ./{} >& {}".format(bdir,code,scrout)
          if ename=='run_schism': scmd="module unload python3;"+scmd 
       elif qnode in ['x5672','vortex','vortexa','c18x','potomac','james','bora']:
          scmd="mvp2run -v -e job_on_node=1 -e bdir='{}' ./{} >& {}".format(bdir,code,scrout)
          if qnode=='bora': scmd="mvp2run -v -a -e job_on_node=1 -e bdir='{}' ./{} >& {}".format(bdir,code,scrout)
          #if qnode=='james' and ename=='run_schism': scmd='mpiexec -np {} --bind-to socket {}/{} >& {}'.format(nproc,bdir,code,scrout)
       else:
          sys.exit('unknow qnode: {},tag=2'.format(qnode))

    return scmd

def compute_contour(x,y,z,levels,fname=None,prj='epsg:4326',show_contour=False,nx=5000,ny=5000):
    '''
    compute contour lines
    Input:
        x: array for x coordinates (ndx)
        y: array for y coordinates (ndy)
        z: matrix of data (ndy,ndx)
        levels: values of contour lines to be extracted
        (nx,ny): when ndx(ndy)>nx(ny), subdivide the domain to speed up
    Output:
        fname: when fname is not None, write contours in shapefiles
        prj:  projection names of shapefiles
        show_contour: plot contours
    '''

    #check level
    fpn=~isnan(z); zmin=z[fpn].min(); zmax=z[fpn].max()
    levels=array(levels); fpz=(levels>=zmin)*(levels<=zmax); levels=sort(levels[fpz])

    #data capsule
    S=zdata(); S.levels=levels; S.xy=[[] for i in arange(len(levels))]
    if len(levels)==0: return S

    #divide domain
    ndx=len(x); ndy=len(y)
    ixs=[]; i1=0; i2=min(nx,ndx)
    while True:
        ixs.append([i1,i2])
        if i2>=ndx: break
        i1=i1+nx; i2=min(i2+nx,ndx)

    iys=[]; i1=0; i2=min(ny,ndy)
    while True:
        iys.append([i1,i2])
        if i2>=ndy: break
        i1=i1+ny; i2=min(i2+ny,ndy)

    #extract contours for subdomains
    for m,[ix1,ix2] in enumerate(ixs):
        for n,[iy1,iy2] in enumerate(iys):
            sxi=x[ix1:ix2]; syi=y[iy1:iy2]; szi=z[iy1:iy2,ix1:ix2]
            fpn=~isnan(szi); zmin=szi[fpn].min(); zmax=szi[fpn].max()
            fpz=(levels>=zmin)*(levels<=zmax); levels_sub=sort(levels[fpz])
            if len(levels_sub)==0: continue
            print('extracting contours in subdomain: {}/{}'.format(n+1+m*len(iys),len(ixs)*len(iys)))

            hf=figure(); hf.set_visible(False)
            P=contour(sxi,syi,szi,levels_sub)
            close(hf)

            for k in arange(len(P.collections)):
                p=P.collections[k].get_paths()
                for i in arange(len(p)):
                    xii=p[i].vertices[:,0]; yii=p[i].vertices[:,1];
                    if i==0:
                        xi=r_[xii,NaN];
                        yi=r_[yii,NaN];
                    else:
                        xi=r_[xi,xii,NaN];
                        yi=r_[yi,yii,NaN];
                #collect contour in each subdomain
                sindc=nonzero(levels==levels_sub[k])[0][0]
                S.xy[sindc].extend(c_[xi,yi])
    for i in arange(len(levels)): S.xy[i]=array(S.xy[i])

    #write contours
    if fname is not None:
        for i,vi in enumerate(levels):
            c=zdata()
            c.type='POLYLINE'
            c.xy=S.xy[i]
            c.prj=get_prj_file(prj)
            cname='{}_{}'.format(fname,vi) if vi>0 else '{}_m{}'.format(fname,-vi)
            write_shapefile_data(cname,c)

    #plot contours
    cs='krgbmcy'
    if show_contour:
        figure()
        for i,vi in enumerate(levels):
            xi,yi=S.xy[i].T
            plot(xi,yi,color=cs[i%7],lw=0.5)
        legend([*levels])

    return S

def load_bathymetry(x,y,fname,z=None,fmt=0):
    '''
    load bathymetry data onto points(xy)
    Input:
        fname: name of DEM file (format can be *.asc or *.npz)

    Outpt:
        fmt=0: return dp; depth interpolated for all points
        fmt=1: return [dpi, sindi]; depth for points only modified, and also flag
    '''

    #input
    xi0=x; yi0=y

    #read DEM
    if fname.endswith('npz'):
        S=loadz(fname,svars=['lon','lat'])
        dx=abs(diff(S.lon)).mean(); dy=abs(diff(S.lat)).mean()
    elif fname.endswith('asc'):
        S=zdata();
        #read *.asc data
        fid=open(fname,'r');
        ncols=int(fid.readline().strip().split()[1])
        nrows=int(fid.readline().strip().split()[1])
        xn,xll=fid.readline().strip().split(); xll=float(xll)
        yn,yll=fid.readline().strip().split(); yll=float(yll)
        dxy=float(fid.readline().strip().split()[1]); dx=dxy; dy=dxy
        nodata=float(fid.readline().strip().split()[1])
        fid.close()

        #shift half a cell if ll defined at corner
        #if xn.lower()=='xllcenter' and yn.lower()=='yllcenter': xll=xll+dxy/2; yll=yll+dxy/2;
        if xn.lower()=='xllcorner' and yn.lower()=='yllcorner': xll=xll+dxy/2; yll=yll+dxy/2

        S.lon=xll+dxy*arange(ncols); S.lat=yll-dxy*arange(nrows)+(nrows-1)*dxy
        S.nodata=nodata
    else:
        sys.exit('wrong format of DEM')

    #check domain of DEM
    if xi0.min()>=(S.lon.max()+dx/2) or xi0.max()<=(S.lon.min()-dx/2) or \
       yi0.min()>=(S.lat.max()+dy/2) or yi0.max()<=(S.lat.min()-dy/2):
       #return depth
       if fmt==0:
          if z is None: z=zeros(len(x))*nan
          return z
       elif fmt==1:
          return [array([]),array([]).astype('int')]
       else:
          sys.exit('wrong fmt')

    #load DEMs
    if fname.endswith('npz'):
        S=loadz(fname)
        if not hasattr(S,'nodata'): S.nodata=None
    elif fname.endswith('asc'):
        S.elev=loadtxt(fname,skiprows=6)

    #change y direction
    if mean(diff(S.lat))<0: S.lat=flipud(S.lat); S.elev=flipud(S.elev)
    lon=S.lon; dx=diff(lon).mean()
    lat=S.lat; dy=diff(lat).mean()
    lon1=lon.min(); lon2=lon.max(); lat1=lat.min(); lat2=lat.max()

    #move (x,y) by half cell
    fpn=(xi0>=(lon1-dx/2))*(xi0<lon1); xi0[fpn]=lon1
    fpn=(yi0>=(lat1-dy/2))*(yi0<lat1); yi0[fpn]=lat1
    fpn=(xi0>=lon2)*(xi0<=(lon2+dx/2)); xi0[fpn]=lon2-dx*1e-6
    fpn=(yi0>=lat2)*(yi0<=(lat2+dy/2)); yi0[fpn]=lat2-dy*1e-6

    #get (x,y) inside dem domain
    sindp=nonzero((xi0>=lon1)*(xi0<=lon2)*(yi0>=lat1)*(yi0<=lat2))[0]
    xi=xi0[sindp]; yi=yi0[sindp]

    #compute index of (x,y)
    idx=floor((xi-lon[0])/dx).astype('int')
    idy=floor((yi-lat[0])/dy).astype('int')

    #make sure lon[idx]<=xi
    sind=nonzero((lon[idx]-xi)>0)[0]
    while len(sind)!=0:
        idx[sind]=idx[sind]-1
        fps=nonzero((lon[idx[sind]]-xi[sind])>0)[0]
        sind=sind[fps]

    #make sure lat[idy]<=yi
    sind=nonzero((lat[idy]-yi)>0)[0]
    while len(sind)!=0:
        idy[sind]=idy[sind]-1
        fps=nonzero((lat[idy[sind]]-yi[sind])>0)[0]
        sind=sind[fps]

    #compute xrat and yrat
    xrat=(xi-lon[idx])/(lon[idx+1]-S.lon[idx])
    yrat=(yi-lat[idy])/(lat[idy+1]-S.lat[idy])
    if sum((xrat<0)*(xrat>1))!=0: sys.exit('xrat<0 or xrat>1')
    if sum((yrat<0)*(yrat>1))!=0: sys.exit('yrat<0 or yrat>1')

    #make sure elevation is within right range
    if S.nodata is not None:
        if isnan(S.nodata):
           fpz=isnan(S.elev)
        else:
           #fpz=(abs(S.elev)>1.5e4)|(S.elev==S.nodata)|(abs(S.elev-S.nodata)<1)
           fpz=S.elev==S.nodata
        S.elev[fpz]=nan

    #compute depth
    dp1=S.elev[idy,idx]*(1-xrat)+S.elev[idy,idx+1]*xrat
    dp2=S.elev[idy+1,idx]*(1-xrat)+S.elev[idy+1,idx+1]*xrat
    dp=dp1*(1-yrat)+dp2*yrat

    #make sure elevation is within right range
    if S.nodata is not None:
        fpz=~isnan(dp); sindp=sindp[fpz]; dp=dp[fpz]

    #return depth
    if fmt==0:
        if z is None: z=zeros(len(x))*nan
        z[sindp]=dp
        return z
    elif fmt==1:
        return [dp, sindp]
    else:
        sys.exit('wrong fmt')

def rewrite_input(fname,qnode=None,nnode=1,ppn=1,**args):
    '''
    function to rewrite the inputs in job-submit scripts (e.g.run_mpi_template.py)
    '''
    #write qnode,nnode,ppn
    if qnode is None: sys.exit('please specify qnode')
    rewrite(fname,replace=['qnode','#qnode'],startswith=['qnode='])
    rewrite(fname,replace=["qnode='{}'; nnode={}; ppn={}\n".format(qnode,nnode,ppn)],startswith=["#qnode='{}'".format(qnode)],note_delimiter='#')       

    #change parameters
    for key,value in args.items(): 
        if key in ['icmb','ifs','fmt','stacks']: 
           fstr="{}={}".format(key,value)
        else:
           fstr="{}='{}'".format(key,value)
        rewrite(fname,replace=[fstr],startswith=['{}='.format(key)],note_delimiter='#')

def rewrite(fname,replace=None,include=None,startswith=None,endswith=None,append=None,note_delimiter=None):
    '''
    function to rewrite file in-situ based on conditions
         fname: file name
         replace: string pairs list; e.g. replace=['ON', 'OFF']
         include: list of strings included; e.g. include=['USE_ICM']
         startswith: list of  strings startswith; e.g. startswith=['#USE_ICM']
         endswith: list of strings endswith;   e.g. endwith=['*.csv']
         append: list of lines; e.g. ['add 1st line','add 2nd line']
         note_delimiter: keep inline note after delimiter (ignore delimiters in the beginning)
    '''

    #read fname
    if os.path.exists(fname):
       fid=open(fname,'r'); lines=fid.readlines(); fid.close()
    else:
       return

    #rewrite lines
    slines=[]
    for line in lines:
        sline=line; iflag=0

        #check include
        if include is not None:
           for i in include:
               if i in sline: iflag=1

        #check startswith
        if startswith is not None:
           for i in startswith:
               if sline.strip().startswith(i): iflag=1

        #check startswith
        if endswith is not None:
           for i in endswith:
               if sline.strip().endswith(i): iflag=1

        #check note
        nd=note_delimiter; note=''
        if (nd is not None) and iflag==1:
           note=sline.strip()
           while note.startswith(nd): note=note[1:]
           if nd in note:
              sid=note.find(nd); note=note[sid:]
           else:
              note=''

        #replace string
        if iflag==1:
           if replace is not None:
              if len(replace)==0:
                 continue
              elif len(replace)==1:
                 sline=replace[0].rstrip()+' '+note
              else:
                 sline=sline.replace(*replace)
           else:
              continue

        #save new line
        if not sline.endswith('\n'): sline=sline+'\n'
        slines.append(sline)

    #append
    if append is not None:
       for sline in append:
           if not sline.endswith('\n'): sline=sline+'\n'
           slines.append(sline)

    #write new line
    fid=open(fname,'w+'); fid.writelines(slines); fid.close()

def convert_dem_format(fname,sname,fmt=0):
    '''
    fname: name of source DEM file
    sname: name of file to be saved
    fmt=0: convert DEM file in *.asc format to *.npz format
    '''

    if fmt==0:
        if not fname.endswith('.asc'): fname=fname+'.asc'

        #read file
        fid=open(fname,'r');
        ncols=int(fid.readline().strip().split()[1])
        nrows=int(fid.readline().strip().split()[1])
        xn,xll=fid.readline().strip().split(); xll=float(xll)
        yn,yll=fid.readline().strip().split(); yll=float(yll)
        dxy=float(fid.readline().strip().split()[1])
        nodata=float(fid.readline().strip().split()[1])
        elev=loadtxt(fname,skiprows=6)
        fid.close()

        #shift half a cell if ll defined at corner
        #if xn.lower()=='xllcenter' and yn.lower()=='yllcenter': xll=xll+dxy/2; yll=yll+dxy/2;
        if xn.lower()=='xllcorner' and yn.lower()=='yllcorner': xll=xll+dxy/2; yll=yll+dxy/2;

        #save data
        S=zdata()
        S.lon=xll+dxy*arange(ncols); S.lat=yll-dxy*arange(nrows)+(nrows-1)*dxy
        S.elev=elev.astype('float32'); S.nodata=nodata
        savez(sname,S)

def plot_taylor_diagram(R=None,STD=None,std_max=2,ticks_R=None,ticks_STD=None,ticks_RMSD=None,
                        cR='b',cSTD='k',cRMSD='g',lw_inner=0.4,lw_outer=2,npt=200,labels=None):
    '''
    plot taylor diagram, and return handles

    Input:
        R: correlation coefficient
        STD: normalized standard dievaiton (or standard dievation)
        std_max: limit of std axis
        ticks_R, ticks_STD, ticks_RMSD: ticks for R, STD, and RMSD
        cR,cSTD,cRMSD: colors for R, STD, and RMSD
        lw_inner, lw_outer: line widths for inner and outer lines
        npt: number of pts for lines
        labels: when labels!=None, add legend

    note: after changing markers' properties, call self.hl.legend() to update legends
    '''

    #get default value for axis
    if ticks_R is None: ticks_R=array([*arange(0.1,1.0,0.1),0.95,0.99])
    if ticks_STD is None: ticks_STD=arange(0.5,5,0.5)
    if ticks_RMSD is None: ticks_RMSD=arange(0.5,5,0.5)
    sm=std_max; S=zdata()

    #plot axis R
    xi=linspace(0,sm,npt)
    S.hp_R=[plot(xi*i,xi*sqrt(1-i**2),ls='--',lw=lw_inner,color=cR) for i in ticks_R]
    S.ht_R=text(0.97*sm*cos(45*pi/180),0.97*sm*sin(45*pi/180),'correlation',fontsize=10,fontweight='bold',color=cR,rotation=-45)
    S.ht_R2=[text(1.01*sm*i,1.01*sm*sqrt(1-i**2), '{}'.format(float(int(i*100))/100),fontsize=8,color=cR) for i in ticks_R]

    #plot STD
    ri=linspace(0,pi/2,npt);
    S.hp_STD=[plot(cos(ri)*i,sin(ri)*i,ls='--',lw=lw_inner,color=cSTD) for i in ticks_STD if i<=sm]
    S.hp_STD2=plot(r_[cos(ri),0,0,1]*sm,r_[sin(ri),1,0,0]*sm,ls='-',lw=lw_outer,color=cSTD)
    S.ht_STD=text(-0.1*sm,0.35*sm,'standard deviation',fontsize=10,fontweight='bold',color=cSTD,rotation=90)

    #plot RMSD
    ri=linspace(0,pi,npt); xi=cos(ri); yi=sin(ri)
    S.hp_RMSD=[]
    for i in ticks_RMSD:
        #line
        xii=xi*i+1; yii=yi*i; fpn=(sqrt(xii**2+yii**2)<sm)*(xii>=0)
        if sum(fpn)==0: continue
        hl=plot(xii[fpn],yii[fpn],ls='--',lw=lw_inner,color=cRMSD)

        #text
        xiii=abs(xii-(sm-0.85*i)/sm); sid=nonzero(xiii==min(xiii))[0][0]
        text(1.02*xii[sid],1.02*yii[sid],'{}'.format(i),color=cRMSD,fontsize=8,rotation=15)

        S.hp_RMSD.append(hl)
    S.ht_RMSD=text(0.08*sm,0.88*sm,'RMSD',color=cRMSD,fontsize=10,fontweight='bold',rotation=25)

    #plot pts
    if (R is not None) and (STD is not None):
        S.hp_obs=plot(0,1,'k.',ms=12,color='k',label='obs')
        S.hp=[];
        for i,ri in enumerate(R):
            xi=ri*STD[i]; yi=sqrt(1-ri**2)*STD[i]
            hp=plot(xi,yi,'r.',ms=10,color='k',label='{}'.format(i))
            S.hp.append(hp)

    #note
    setp(gca(),yticks=ticks_STD,xticks=[]); yticks(fontsize=8)
    gca().spines['right'].set_visible(False)
    gca().spines['top'].set_visible(False)
    gca().spines['left'].set_visible(False)
    gca().spines['bottom'].set_visible(False)
    df=1e-3; setp(gca(),xlim=[-df,sm+df],ylim=[-df,sm+df])
    S.ha=gca(); S.ax=gca().axes;

    def update_legend(self=S,**args):
        self.hl=self.ha.legend(**args)
    S.update_legend=update_legend

    #add legend
    if labels is not None:
        S.hl=S.ha.legend(fontsize=8)
        if hasattr(labels,'__len__'): [S.hl.get_texts()[i+1].set_text(lstr) for i,lstr in enumerate(labels)]

    return S
def get_subplot_position2(margin=[0.1,0.1,0.1,0.1],dxy=[0.05,0.05],ds=[3,4],**args):
    '''
    return subplot position. Based on and calling get_subplot_position, but using the margin as input. 
    Inut:
        margin=[left,right,up,down]: distance from left, right, up, and bottom edge
        for other arguments, see get_subplot_position
    Sample function call:
        [ps,pc]=get_subplot_position2(margin=[0.05,0.05,0.1,0.1],dxy=[0.00,0.00],ds=[3,4],dc=[0.01,0.005])
        ps=reshape(ps,(12,4)) #to make 3D dimension array to 2 dimension
        for imon in arange(12)+1:
            axes(position=ps[imon-1])
    '''
    left,right,up,down=margin
    rown,coln=ds
    xspan=(1-left-right-(coln-1)*dxy[0])/coln #get x-span for each subplot
    yspan=(1-up-down-(rown-1)*dxy[1])/rown #get y-span for each subplot
    p0=[left,1-up-yspan,xspan,yspan] #get the upper left subplot position
    return get_subplot_position(p0,dxy,ds,**args)

def get_subplot_position(p0,dxy,ds,dc=None,sindc=None,figsize=None):
    '''
    return subplot position
    Input:
       p0=[x0,y0,xm,ym]: upper left subplot position
       dxy=[dx,dy]:      space between subplots
       ds=[ny,nx]:       subplot structure
       dc=[xmc,dxc]:     add colorbar position with width of xmc, and distance dxc from axes
       sindc=[:nplot]:   indices of subplot colorbars
       fsize=[fw,fh]:    plot out subplot in figure(figsize=fsize)
    '''

    #compute subplot position
    x0,y0,xm,ym=p0; dx,dy=dxy; ny,nx=ds
    ps=array([[[x0+i*(xm+dx),y0-k*(ym+dy),xm,ym] for i in arange(nx)] for k in arange(ny)])
    if dc!=None:
       xmc,dxc=dc; pc=zeros(ds).astype('O');  pc[:]=0
       if sindc!=None: pc.ravel()[setdiff1d(arange(prod(ds)),array(sindc))]=0
       for k in arange(ny):
           for i in arange(nx):
               if pc[k,i]!=None: pc[k,i]=[x0+xm+i*(xm+dx)+dxc,y0-k*(ym+dy),xmc,ym]

    #plot subplots
    if figsize!=None:
       figure(figsize=figsize)
       for i in arange(nx):
           for k in arange(ny):
               axes(position=ps[k,i]); xticks([]); yticks([])
               #setp(gca(),xticklabels=[],yticklabels=[])
               if dc!=None:
                  if pc[k,i]!=None: axes(position=pc[k,i]); xticks([]); yticks([])
    if dc!=None:
       return [ps,pc]
    else:
       return ps

def close_data_loop(xi):
    '''
    constructe data loop along the first dimension.
    if xi[0,...]!=xi[-1,...], then,add xi[0,...] in the end
    '''
    if array_equal(xi[0,...].ravel(),xi[-1,...].ravel()):
        vi=xi
    else:
        vi=r_[xi,xi[0,...][None,...]]
    return vi

# def close_data_loop(xi):
#     '''
#     constructe data loop along the last dimension.
#     if xi[...,:,0]!=xi[...,:,-1], then,add xi[...,:,0] in the end
#     '''
#     if xi.ndim==1:
#        if xi[0]!=xi[-1]:
#           vi=r_[xi,xi[0]];
#        else:
#           vi=xi
#     else:
#        if array_equal(xi[...,:,0].ravel(),xi[...,:,-1].ravel()):
#           vi=xi
#        else:
#           vi=c_[xi,xi[...,:,0][...,None]]
#     return vi

def find_cs(xi,dx):
    '''
    analyze time series to find the locations where differences are larger than dx: (xi[i+1]-xi[i]>dx)
    return: 
      bind: locations of differences>dx
      sections: continous sections where all differences are smaller than dx 
      gaps:     gap sections where differences are larger than dx 
      slen,glen: lens for each section or gap 
      msection,mgap: maximum section/gap
    '''
    sind=nonzero(diff(xi)>dx)[0]; sections=[]; gaps,glen,mgap=[],[],[]
    if len(sind)==0:
       sx=[xi[0],xi[-1]]; sections.append(sx)
    else:
       sx=[xi[0],xi[sind[0]]]; sections.append(sx)
       for m in arange(len(sind)-1):
           sx=[xi[sind[m]+1],xi[sind[m+1]]]; sections.append(sx) 
           gx=[xi[sind[m]],xi[sind[m]+1]]; gaps.append(gx)
       sx=[xi[sind[-1]+1],xi[-1]]; sections.append(sx)
    sections=array(sections); gaps=array(gaps)
    slen=diff(sections,axis=1); msection=sections[nonzero(slen==slen.max())[0][0]]
    if len(gaps)!=0: glen=diff(gaps,axis=1); mgap=gaps[nonzero(glen==glen.max())[0][0]]
    S=zdata(); S.bind,S.sections,S.msection,S.slen,S.gaps,S.mgap,S.glen=sind,sections,msection,slen,gaps,mgap,glen
    return S

def datenum(*args,fmt=0):
    '''
    usage: datenum(*args,fmt=[0,1])
       datenum(2001,1,1,10,23,0)       or datenum([[2001,1,1],[2002,1,5]])
       datenum('2001-01-01, 10:23:00') or datenum(['2001-01-1','2002-01-05'])
       fmt=0: output num; fmt==1: output date
    '''

    #input only one argument, it should be a time string
    if len(args)==1: args=args[0]

    if isinstance(args,str):
       #single time string
       dnum=datestr2num(args)
       if fmt!=0: dnum=num2date(dnum)
    elif hasattr(args,"__len__"):
       if isinstance(args[0],str)|hasattr(args[0],"__len__"):
          #array of time string or time array
          dnum=array([datenum(i,fmt=fmt) for i in args])
       else:
          dargs=[*args]
          #if month >12: convert
          if dargs[1]>12:
             dargs[0]=dargs[0]+int(dargs[1]/12)
             dargs[1]=max(1,dargs[1]%12)

          #time array (e.g. [2002,1,1])
          dnum=datetime.datetime(*dargs)
          if fmt==0: dnum=mpl.dates.date2num(dnum)
    else:
       sys.exit('unknown input format')

    return dnum

def get_xtick(fmt=0,xts=None,str=None):
    '''
    return temporal ticks and labels for plot purpose

        fmt: format of xtick and xticklabel
             0: year; 1: month; 2: day;  3: user-defined
        xts: time aranges or time arrays
            e.g. [2000,2010],[datenum(2000,1,1),datenum(2010,1,1)],[*arange(730120,730420)]
        str: format of label
            Year:   %y=01;  %-y=1;   %Y=2000
            Month:  %m=01;  %-m=1;   %B=January;  %b=Jan
            Day:    %d=01;  %-d=1
            Hour:   %H=09;  %-H=9;   %I=[00,12];  %-I=1
            Minute: %M=09;  %-M=9;
            Second: %S=09;  %-S=9
            AM/PM:  %p=[AM,PM]

            %c='Fri Jan 25 04:05:02 2008'
            %x='01/25/08'
            %X='04:05:02'

            Week:           %a=MON;      %A=Monday;   %w=[0,6]
            Week of year:   %U=[00,53];  %W=[00,53]
            Day of year:    %j=045;       %-j=45

            1: 2008-02-03
            2: 2008-02-03, 04:05:00
            3: J,F,M,A,J,J,... (Months)
            4： 03/15 (mm/dd)
    '''
    #get time label
    ft=0
    if str==1: str='%Y-%m-%d'
    if str==2: str='%Y-%m-%d, %H:%M:%S'
    if str==3: str='%b'; ft=1
    if str==4: str='%m/%d'

    #get time ticks
    it=0
    if xts is None:
        if fmt==3: sys.exit('must provide xts for fmt=3')
        if fmt==2:
            xts=[2000,2000]
        else:
            xts=[*arange(2000,2025)]
    elif len(xts)==2: #[year1,year2] or [datenum1, datenum2]
        if xts[0]>1e4: it=1
        xts=[*arange(xts[0],xts[1]+1)]
    else: #array of datenum
        it=1

    #compute time ticks and ticklabels
    if fmt==0:   #for year
        if str is None: str='%Y'
        if it==0: xts=[datenum(i,1,1) for i in xts]
    elif fmt==1: #for months
        if str is None: str='%b'
        if it==0: xts=array([[datenum(i,k+1,1) for k in arange(12)] for i in xts]).ravel()
    elif fmt==2: #for days
        if str is None: str='%-d'
        if it==0: xts=[*arange(datenum(min(xts),1,1),datenum(max(xts)+1,1,1)-1)]
    elif fmt==3: #user defined
        if str is None: str='%-d'
    xls=[num2date(i).strftime(str) for i in xts]
    if ft==1: xls=[i[0] for i in xls]

    return [xts,xls]

#-------loadz------------------------------------------------------------------
class zdata:
    '''
    self-defined data structure by Zhengui Wang.  Attributes are used to store data
    '''
    def __init__(self):
        pass

def savez(fname,data,fmt=0):
    '''
    save data as self-defined python format
       fmt=0: save data as *.npz (small filesize, reads slower)
       fmt=1: save data as *.pkl (large filesize, reads faster)
    if fname endswith *.npz or *.pkl, then fmt is reset to match fname
    '''

    #determine format
    if fname.endswith('.npz'): fmt=0; fname=fname[:-4]
    if fname.endswith('.pkl'): fmt=1; fname=fname[:-4]
    if fmt==1: fname=fname+'.pkl'

    #save data
    if fmt==0:
       #get all attribute
       svars=list(data.__dict__.keys())
       if 'VINFO' in svars: svars.remove('VINFO')

       #check whether there are functions. If yes, change function to string
       rvars=[]
       for vari in svars:
           if hasattr(data.__dict__[vari], '__call__'):
              import cloudpickle
              try:
                 data.__dict__[vari]=cloudpickle.dumps(data.__dict__[vari])
              except:
                 print('function {} not saved'.format(vari))
                 rvars.append(vari)
       svars=setdiff1d(svars,rvars)

       #constrcut save_string
       save_str='savez_compressed("{}" '.format(fname)
       for vari in svars: save_str=save_str+',{}=data.{}'.format(vari,vari)
       save_str=save_str+')';  exec(save_str)
    elif fmt==1:
       if hasattr(data,'VINFO'): del data.VINFO
       fid=open(fname,'wb'); pickle.dump(data,fid,pickle.HIGHEST_PROTOCOL); fid.close()

def loadz(fname,svars=None):
    '''
    load self-defined data "fname.npz" or "fname.pkl"
         svars: list of variables to be read
    '''

    if fname.endswith('.npz'):
       #get data info
       data0=load(fname,allow_pickle=True)
       keys0=data0.keys() if svars is None else svars

       #extract data, and VINFO is used to store data info
       vdata=zdata();  VINFO=[]
       for keyi in keys0:
           datai=data0[keyi];
           #if value is a object
           if datai.dtype==dtype('O'): datai=datai[()]

           #if value is a function
           if 'cloudpickle.cloudpickle' in str(datai):
              import pickle
              try:
                 datai=pickle.loads(datai)
              except:
                 continue

           #output format
           exec('vdata.'+keyi+'=datai')
           ##gather information about datai
           #vinfo=keyi+": "+type(datai).__name__
           #if isinstance(datai,list):
           #    vinfo=vinfo+'('+str(len(datai))+'), '
           #elif isinstance(datai,np.ndarray):
           #    vinfo=vinfo+str(datai.shape)+', dtype='+str(datai.dtype)
           #VINFO.append(vinfo)
       #vdata.VINFO=array(VINFO)
    elif fname.endswith('.pkl'):
       import pickle
       vdata=zdata(); fid=open(fname,'rb')
       data=pickle.load(fid)
       vdata.__dict__=dcopy(data).__dict__.copy()
       fid.close()
    else:
       sys.exit('unknown format: {}'.format(fname))

    #gather vdata information
    fs=[]
    for i in vdata.__dict__.keys():
        vi=vdata.__dict__[i]; f0='{}: '.format(i)
        if isinstance(vi,list):
           f1='{}{}'.format(str(type(vi)),len(vi))
        elif isinstance(vi,np.ndarray):
           f1='ndarray{}'.format(vi.shape)
        else:
           f1='{}'.format(type(vi))
        f2=' ,dtype={}'.format(str(vi.dtype)) if hasattr(vi,'dtype') else ''
        fs.append(f0+f1+f2)
    vdata.VINFO=array(fs)

    return vdata

def least_square_fit(X,Y):
    '''
    perform least square fit
    usage: CC,fit_data=least_square_fit(X,Y)
        X: data maxtrix (npt,n)
        Y: data to be fitted (npt)
    where CC(n) is coefficient, fit_data is data fitted
    '''
    if X.shape[0]!=len(Y): X=X.T
    CC=inv(X.T@X)@(X.T@Y); fy=X@CC

    return [CC,fy]

#-------mfft-------------------------------------------------------------------
def mfft(xi,dt):
    '''
    Perform FFT for a time series, with a time interval specified

    usage: period,afx,pfx=mfft(xi,dt)
    input:
       xi: time series
       dt: time interval

    output:
       period[period],afx[amplitude],pfx[phase]
    '''
    N=xi.size;
    fx=fft(xi);
    afx=abs(fx[1:N//2])*2.0/N;
    pfx=angle(fx[1:N//2]);
    period=dt*N/arange(1,N//2);
    return period,afx,pfx

def command_outputs(code,shell=True):
    '''
    Capture the command output from the system

    usage:
         S=command_outputs('ls')
         print(S.stderr)
         print(S.stdout) # normally this is the results
    '''
    import subprocess
    p=subprocess.Popen(code,stdout=subprocess.PIPE,stderr=subprocess.STDOUT,shell=shell)
    stdout,stderr=p.communicate()
    if stdout!=None: stdout=stdout.decode('utf-8')
    if stderr!=None: stderr=stderr.decode('utf-8')

    out=zdata()
    out.stdout=stdout
    out.stderr=stderr
    return out

def near_pts(pts,pts0,method=0,N=100):
    '''
    return index of pts0 that pts is nearest
       usage: sind=near_pts(pts,pts0)
       pts0: c_[x0,y0];  pts: c_[x,y]
       algorithm: using sp.spatial.cKDTree (default)

    old methods: method=1 and  method=2
       usage: sind=near_pts(pts,pts0,method=1[2],N=100)
       pts[n,2]: xy of points
       pts0[n,2]: xy of points

       method=1: quick method by subgroups (N);
       method=2: slower methods
    '''

    if method==0:
        sind=sp.spatial.cKDTree(pts0).query(pts)[1]
    elif method==1:
        p=pts[:,0]+(1j)*pts[:,1]
        p0=pts0[:,0]+(1j)*pts0[:,1]

        # N=min(N,len(p)-1);
        # #--divide pts into subgroups based on dist, each group size is about N---------
        # ps0=[]; ps=[]; ds=[]; inds=[]
        # i0=0; mval=1e50; pflag=p==None
        # while(True):
        #     ps0.append(p[i0]);
        #     dist=abs(p-p[i0]); dist[pflag]=mval;
        #     ind=argsort(dist); sdist=dist[ind]; mds=sdist[N]; i0=ind[N]
        #     fp=dist<mds; psi=p[fp]; dsi=max(dist[fp]); pflag[fp]=True;
        #     ps.append(psi); ds.append(dsi); inds.append(nonzero(fp)[0])
        #     if mds==mval: break

        N=min(N,len(p));
        #--divide pts into subgroups based on dist, each group size is about N---------
        ps0=[]; ps=[]; ds=[]; inds=[]; inum=arange(len(p))
        while(True):
            if len(inum)==0: break
            dist=abs(p[inum]-p[inum[0]]); sind=argsort(dist); inum=inum[sind]; dist=dist[sind]; sN=min(N,len(inum))
            ps0.append(p[inum[0]]); ps.append(p[inum[:sN]]); ds.append(dist[sN-1]); inds.append(inum[:sN])
            inum=inum[sN:]

        #---find nearest pts for each subgroup-----------------------------------------
        inds0=[]
        for m in arange(len(ps)):
            dist=abs(p0-ps0[m]); dsi=ds[m]; psi=ps[m]
            #---find radius around ps0[m]
            dsm=dsi
            while(True):
                fp=dist<=dsm;
                if sum(fp)>0:
                    break;
                else:
                    dsm=dsm+dsi;
            #-----subgroup pts of p0---------------------------------
            fp=dist<=(dsm+2*dsi); ind0=nonzero(fp)[0]; p0i=p0[ind0];
            psii=psi[:,None]; p0ii=p0i[None,:]
            dist=abs(psii-p0ii); indi=dist.argmin(axis=1);
            inds0.append(ind0[indi]);

        #-arrange index-----------------
        pind=array([]).astype('int'); pind0=array([]).astype('int')
        for m in arange(len(inds)):
            pind=r_[pind,inds[m]]
            pind0=r_[pind0,inds0[m]]

        ind=argsort(pind);
        sind=pind0[ind]
    elif method==2:
        n=pts.shape[0]; n0=pts0.shape[0]
        N=max(min(1e7//n0,n),1e2)
        print('total pts: {}'.format(n));

        i0=int(0); i1=int(N);
        while(True):
            print('processing pts: {}-{}'.format(i0,i1));
            x=pts[i0:i1,0]; y=pts[i0:i1,1]
            x0=pts0[:,0]; y0=pts0[:,1]
            dist=(x[None,:]-x0[:,None])**2+(y[None,:]-y0[:,None])**2

            indi=[];
            for i in arange(x.shape[0]):
                disti=dist[:,i];
                indi.append(nonzero(disti==min(disti))[0][0])

            if i0==0:
                ind=array(indi);
            else:
                ind=r_[ind,squeeze(array(indi))];
            #next step
            i0=int(i0+N); i1=int(min(i1+N,n))
            if i0>=n: break
        sind=ind

    return sind

def inside_polygon(pts,px,py,fmt=0,method=0):
    '''
    check whether points are inside polygons

    usage: sind=inside_polygon(pts,px,py):
       pts[npt,2]: xy of points
       px[npt] or px[npt,nploy]: x coordiations of polygons
       py[npt] or py[npt,nploy]: y coordiations of polygons
       (npt is number of points, nploy is number of polygons)

       fmt=0: return flags "index[npt,nploy]" for whether points are inside polygons (1 means Yes, 0 means No)
       fmt=1: only return the indices "index[npt]" of polygons that pts resides in
             (if a point is inside multiple polygons, only one indice is returned; -1 mean pt outside of all Polygons)

       method=0: use mpl.path.Path
       method=1: use ray method explicitly

    note: For method=1, the computation time is proportional to npt**2 of the polygons. If the geometry
          of polygons are too complex, dividing them to subregions will increase efficiency.
    '''
    #----use ray method-----------------------------
    #check dimension
    if px.ndim==1:
       px=px[:,None]; py=py[:,None]

    #get dimensions
    npt=pts.shape[0]; nv,npy=px.shape

    if nv==3 and fmt==1:
       #for triangles, and only return indices of polygons that points resides in
       px1=px.min(axis=0); px2=px.max(axis=0); py1=py.min(axis=0); py2=py.max(axis=0)

       sind=[];
       for i in arange(npt):
           pxi=pts[i,0]; pyi=pts[i,1]
           sindp=nonzero((pxi>=px1)*(pxi<=px2)*(pyi>=py1)*(pyi<=py2))[0]; npy=len(sindp)
           if npy==0:
               sind.append(-1)
           else:
               isum=ones(npy)
               for m in arange(nv):
                   xi=c_[ones(npy)*pxi,px[m,sindp],px[mod(m+1,nv),sindp]]
                   yi=c_[ones(npy)*pyi,py[m,sindp],py[mod(m+1,nv),sindp]]
                   area=signa(xi,yi)
                   fp=area<0; isum[fp]=0;
               sindi=nonzero(isum!=0)[0]

               if len(sindi)==0:
                   sind.append(-1)
               else:
                   sind.append(sindp[sindi[0]])
       sind=array(sind)

    else:
        if method==0:
            sind=[]
            for m in arange(npy):
                sindi=mpl.path.Path(c_[px[:,m],py[:,m]]).contains_points(pts)
                sind.append(sindi)
            sind=array(sind).T+0  #convert logical to int
        elif method==1  :
            #using ray method explicitly
            sind=ones([npt,npy])
            x1=pts[:,0][:,None]; y1=pts[:,1][:,None]
            for m in arange(nv):
                x2=px[m,:][None,:]; y2=py[m,:][None,:]; isum=zeros([npt,npy])
                # sign_x1_x2=sign(x1-x2)
                for n in arange(1,nv-1):
                    x3=px[(n+m)%nv,:][None,:]; y3=py[(n+m)%nv,:][None,:]
                    x4=px[(n+m+1)%nv,:][None,:]; y4=py[(n+m+1)%nv,:][None,:]

                    #formulation for a ray to intersect with a line
                    fp1=((y1-y3)*(x2-x1)+(x3-x1)*(y2-y1))*((y1-y4)*(x2-x1)+(x4-x1)*(y2-y1))<=0 #intersection inside line P3P4
                    fp2=((y2-y1)*(x4-x3)-(y4-y3)*(x2-x1))*((y4-y3)*(x1-x3)+(y3-y1)*(x4-x3))<=0 #P1, P2 are in the same side of P3P4

                    fp12=fp1*fp2
                    isum[fp12]=isum[fp12]+1
                fp=((isum%2)==0)|((x1==x2)*(y1==y2))
                sind[fp]=0

        #change format
        if fmt==1:
            sindm=argmax(sind,axis=1)
            sindm[sind[arange(npt),sindm]==0]=-1
            sind=sindm
        elif fmt==0 and npy==1:
            sind=sind[:,0]
    return sind

def signa(x,y):
    '''
        compute signed area for triangles along the last dimension (x[...,0:3],y[...,0:3])
    '''
    if x.ndim==1:
        area=((x[0]-x[2])*(y[1]-y[2])-(x[1]-x[2])*(y[0]-y[2]))/2
    elif x.ndim==2:
        # area=((x[:,0]-x[:,2])*(y[:,1]-y[:,2])-(x[:,1]-x[:,2])*(y[:,0]-y[:,2]))/2
        area=((x[...,0]-x[...,2])*(y[...,1]-y[...,2])-(x[...,1]-x[...,2])*(y[...,0]-y[...,2]))/2
    area=np.squeeze(area)
    return area

def mdivide(A,B):
    '''
        perform matrix division B/A
    '''
    if A.ndim==1: A=A[None,:]
    if B.ndim==1: B=B[None,:]
    A2=inv(A@A.T);
    B2=B@A.T
    return squeeze(B2@A2)

def lpfilt(data,delta_t,cutoff_f):
    '''
    low pass filter for 1D (data[time]) or nD (data[time,...]) array along the first dimension
    '''
    #import gc #discard

    ds=data.shape

    #fft original data
    mdata=data.mean(axis=0)[None,...]
    #print(data.shape,mdata.shape)
    data=data-mdata
    fdata=fft(data,axis=0)

    #desgin filter
    N=ds[0];
    filt=ones(N)
    k=int(floor(cutoff_f*N*delta_t))
    filt[k]=0.715
    filt[k+1]=0.24
    filt[k+2]=0.024
    filt[k+3:N-(k+4)]=0.0
    filt[N-(k+4)]=0.024
    filt[N-(k+3)]=0.24
    filt[N-(k+2)]=0.715

    #expand dimension of filt
    filt=(ones([*ones(data.ndim).astype('int')])*filt).T

    #remove high freqs
    fdata=fdata*filt

    #lp results
    lfdata=real(ifft(fdata,axis=0))+mdata

    return lfdata

def smooth(xi,N):
    '''
    smooth average (on the 1st dimension):
       xi[time,...]: time series
       N: window size (if N is even, then N=N+1)
    '''

    #window span
    if mod(N,2)==0: N=N+1
    nz=int((N-1)/2)

    #moving averaging
    X=xi.copy(); SN=N*ones(xi.shape)
    for nmove in arange(1,nz+1):
        #sum in the beginning and end
        X[nmove:,...]=X[nmove:,...]+xi[:-nmove,...]
        X[:-nmove,...]=X[:-nmove,...]+xi[nmove:,...]

        #count
        SN[:nmove,...]=SN[:nmove,...]-1
        SN[-nmove:]=SN[-nmove:]-1
    SX=X/SN
    return SX

def daytime_length(lat,doy):
    '''
    calculate daytime length based on latitutde and day_of_year
    lat: latitude, doy: (1-365), sunrise=12-daytimelength/2, sunset=12+daytimelength/2
    '''
    P=arcsin(0.39795*cos(0.2163108 + 2*arctan(0.9671396*tan(0.00860*(doy-186)))));
    T=(sin(0.8333*pi/180)+sin(lat*pi/180)*sin(P))/cos(lat*pi/180)/cos(P);
    dt=24-(24/pi)*arccos(T);
    return dt

def move_figure(x=0,y=0,f=None):
    '''
    Move figure (f=gcf()) to upper left corner to pixel (x, y)
    e.g. move_figure(0,0,gcf())
    '''
    if f is None: f=gcf()
    backend = matplotlib.get_backend()
    if backend == 'TkAgg':
        f.canvas.manager.window.wm_geometry("+%d+%d" % (x, y))
    elif backend == 'WXAgg':
        f.canvas.manager.window.SetPosition((x, y))
    else:
        # This works for QT and GTK
        # You can also use window.setGeometry
        f.canvas.manager.window.move(x, y)

def proj(fname0=None,fmt0=None,prj0=None,fname1=None,fmt1=None,prj1=None,x=None,y=None,lon0=None,lat0=None,order0=0,order1=0):
    '''
    tranfrom projection of files: proj(fname0,fmt0,prj0,fname1,fmt1,prj1,x,y,lon0,lat0,order0,order1)
       fname: file name
       fmt: 0: SCHISM gr3 file; 1: SCHISM bp file; 2: xyz file; 3: xyz file with line number
       prj: projection name (e.g. 'epsg:26918', 'epsg:4326','cpp'), or projection string (e.g. prj0=get_prj_file('epsg:4326'))
       lon0,lat0: center for transformation between lon&lat and x&y; if lon0=None or lat0=None, then x0=mean(lon), and y0=mean(lat)
       order=0 for projected coordinate; order=1 for lat&lon (if prj='epsg:4326', order=1 is applied automatically')

    tranform data directly: px,py=proj(prj0='epsg:26918',prj1='epsg:4326',x=px,y=py)

    function used:
          lat,lon=Transformer.from_crs('epsg:26918','epsg:4326').transform(x,y)
          x,y=Transformer.from_crs('epsg:4326','epsg:26918').transform(lat,lon)
    #x1,y1=transform(Proj(proj0),Proj(proj1),x,y); #not used anymore
    '''

    from .schism_file import read_schism_hgrid,read_schism_bpfile,schism_bpfile

    #read file
    if fmt0==0:
        gd=read_schism_hgrid(fname0)
        np=gd.np; x=gd.x; y=gd.y; z=gd.dp
    elif fmt0==1:
        gd=read_schism_bpfile(fname0)
        np=gd.nsta; x=gd.x; y=gd.y; z=gd.z
    elif fmt0==2:
        gd=loadtxt(fname0)
        np=gd.shape[0]; x=gd[:,0]; y=gd[:,1]; z=gd[:,2]
    elif fmt0==3:
        gd=loadtxt(fname0)
        np=gd.shape[0]; x=gd[:,1]; y=gd[:,2]; z=gd[:,3]
    else:
        if x is None or y is None: sys.exit('unknown format of input files: {}, {}'.format(fname0,fname1))

    #order of data
    prj0=prj0.lower(); prj1=prj1.lower(); icpp=0
    if prj0=='epsg:4326': order0=1
    if prj1=='epsg:4326': order1=1
    if 'cpp' in [prj0,prj1]: icpp=1; rearth=6378206.4

    #transform coordinate
    if icpp==1 and prj0=='cpp':
        if prj1!='epsg:4326': sys.exit('projection wrong: prj0={}, prj1={}'.format(prj0,prj1))
        if (lon0 is None) or (lat0 is None): sys.exit('need lon0 and lat0 for cpp=>ll transform')
        x1=lon0+x*180/(pi*rearth*cos(lat0*pi/180)); y1=y*180/(pi*rearth)
    elif icpp==1 and prj1=='cpp':
        if prj0!='epsg:4326': sys.exit('projection wrong: prj0={}, prj1={}'.format(prj0,prj1))
        if lon0 is None: lon0=mean(x)
        if lat0 is None: lat0=mean(y)
        x1=rearth*(x-lon0)*(pi/180)*cos(lat0*pi/180); y1=rearth*y*pi/180
    else:
        if order0==1: x,y=y,x
        fpn=~(isnan(x)|isnan(y)); x1=arange(len(x))*nan; y1=arange(len(y))*nan
        x1[fpn],y1[fpn]=Transformer.from_crs(prj0,prj1).transform(x[fpn],y[fpn])
        if order1==1: x1,y1=y1,x1
        if (sum(isnan(x1[fpn]))!=0) | (sum(isnan(y1[fpn]))!=0): sys.exit('nan found in tranformation: x1,y1') #check nan

    # if (sum(isnan(x))!=0) and (sum(isnan(y))!=0): sys.exit('nan found in x,y') #check nan
    #x1,y1=Transformer.from_crs(prj0,prj1).transform(x,y)
    #x1,y1=transform(Proj(init=proj0),Proj(init=proj1),x,y);
    #x1,y1=transform(Proj(proj0),Proj(proj1),x,y);

    #write file
    if fmt1==0:
        if fmt0!=0: sys.exit('{} should have gr3 format'.format(fname0))
        gd.x=x1; gd.y=y1;
        gd.write_hgrid(fname1,Info='!coordinate transformed: {}=>{}'.format(prj0,prj1))
    elif fmt1==1:
        if fmt0==1:
            gd.x=x1; gd.y=y1
        else:
            gd=schism_bpfile()
            gd.note='coordinate transformed: {}=>{}'.format(prj0,prj1)
            gd.nsta=np; gd.x=x1; gd.y=y1; gd.z=z;
        gd.write_bpfile(fname1)
    elif fmt1==2 or fmt1==3:
        with open(fname1,'w+') as fid:
            for i in arange(np):
                if fmt1==2: fid.write('{} {} {}\n'.format(x1[i],y1[i],z[i]))
                if fmt1==3: fid.write('{} {} {} {}\n'.format(i+1,x1[i],y1[i],z[i]))
    else:
       return [x1,y1]

def proj_pts(x,y,prj1='epsg:4326',prj2='epsg:26918'):
    '''
    convert projection of points from prj1 to prj2
      x,y: coordinate of pts
      prj1: name of original projection
      prj2: name of target projection
    '''
    px,py=proj(prj0=prj1,prj1=prj2,x=x,y=y)
    return [px,py]

def get_prj_file(prjname='epsg:4326',fmt=0,prj_dir=r'D:\Work\Database\projection\prj_files'):
    '''
    return projection name or entire database (dict)
        fmt=0: get one projection
        fmt=1: return dict of projection database
        fmt=-1: process *.prj files from prj_dir

    #-------online method-----------------
    #function to generate .prj file information using spatialreference.org
    #def getWKT_PRJ (epsg_code):
    #     import urllib
    #     # access projection information
    #     wkt = urllib.urlopen("http://spatialreference.org/ref/epsg/{0}/prettywkt/".format(epsg_code))
    #     # remove spaces between charachters
    #     remove_spaces = wkt.read().replace(" ","")
    #     # place all the text on one line
    #     output = remove_spaces.replace("\n", "")
    #     return output
    '''

    #get location of database
    bdir=os.path.dirname(__file__)

    if fmt==0:
        S=loadz('{}/prj.npz'.format(bdir))
        return S.prj[prjname]
    elif fmt==1:
        S=loadz('{}/prj.npz'.format(bdir))
        return S.prj
    elif fmt==-1:
        #dir of *.prj files
        #prj_dir=r'D:\Work\Database\projection\prj_files'

        #---processing all prj files-------------------------------------------
        fnames=os.listdir(prj_dir)

        prj=dict()
        for fname in fnames:
            R=re.match('epsg.(\d+).prj',fname);
            if not R: continue
            prj_num=int(R.groups()[0])
            with open('{}/{}'.format(prj_dir,fname),'r') as fid:
                line=fid.readline().strip();
                prj['epsg:{}'.format(prj_num)]=line

        # #save prj file as a database
        # S=zdata();
        # S.prj=prj
        # savez('{}/prj'.format(bdir),S)
        return prj
    else:
        sys.exit('unknow fmt')

#----------------convert_matfile_format----------------------------------------
#def convert_matfile_format(file):
#    '''
#    convert a matlab data to self-defined python format
#    file: input, a directory or a matfile
#    '''
#    # eg. file="C:\Users\Zhengui\Desktop\Observation2\DWR\SFBay_DWRData_SSI.mat"
#    # file="D:\OneDrive\Python\tem.mat"
#
#    fname=[]
#    if os.path.isdir(file):
#        sfile=os.listdir(file)
#        for sfilei in sfile:
#            ename=sfilei.rstrip().split('.')[-1]
#            if ename=='mat':
#                fname.append(file+os.sep+sfilei)
#    else:
#        fname=[file];
#
#    fid=open('log_matfile_convert.txt','w+')
#    #convert mat format from v7.3 to v7
#    cdir=os.getcwd();
#    os.chdir('D:\OneDrive\Python')
#    import matlab.engine
#    eng = matlab.engine.start_matlab()
#
#    for fn in fname:
#        print('converting matfile: '+fn)
#        dname=os.path.dirname(fn)
#        bname=os.path.basename(fn).split('.')[0]
#        fnv7=dname+os.sep+bname+'_v7'
#        fnz=dname+os.sep+bname
#        eflag=eng.convert_matfile_format(fn,fnv7,nargout=1)
#        if eflag!=0:
#            print('convert flag is %d: %s\n' % (eflag, fn));
#            fid.write('convert flag is %d: %s\n' % (eflag,fn))
#            continue
#        convert_matfile(fnz,fnv7)
#        os.remove(fnv7+'.mat')
#    fid.close()
#    os.chdir(cdir)
#

#convert MATLAB matfile to zdata
def convert_matfile(name_matfile,name_save=None):
    '''
    convert MATLAB file to zdata
      examples:
           1. convert_matfile('wqdata.mat')
           2. convert_matfile('wqdata')
           3. convert_matfile('wqdata','sfdata')
    '''
    from scipy import io

    #check name
    if name_matfile.endswith('.mat'): name_matfile[:-4]
    if name_save is None: name_save=name_matfile

    #read matfile and convert
    C=sp.io.loadmat(name_matfile+'.mat'); S=zdata()
    for keyi in C.keys():
        if keyi.startswith('__'):continue
        valuei=squeeze(C[keyi])

        #change format
        if not issubdtype(valuei.dtype,np.number):
            valuei=array([i[0] if len(i)!=0 else '' for i in valuei])
        if keyi in ['Doy', 'doy']: valuei=valuei-366
        exec('S.{}=squeeze(valuei)'.format(keyi))
    savez(name_save,S)

def get_stat(xi_model,xi_obs,fmt=0):
    '''
    compute statistics between two time series
    x1, x2 must have the same dimension
    x1: model; x2: obs
    fmt=1: compute pvalue using scipy.stats.pearsonr

    #import matlab.engine
    #eng=matlab.engine.start_matlab()
    '''

    x1=xi_model; x2=xi_obs
    mx1=mean(x1); mx2=mean(x2)

    S=zdata(); dx=x1-x2; std1=std(x1); std2=std(x2)
    #---save data
    S.R=corrcoef(x1,x2)[0,1] #R
    S.ME=mean(dx)
    S.MAE=mean(abs(dx))
    S.RMSD=sqrt((dx**2).mean())
    S.std=std(dx)
    S.ms=1-sum(dx**2)/sum((abs(x1-mx2)+abs(x2-mx2))**2)
    if fmt==1: a,S.pvalue=sp.stats.pearsonr(x1,x2)
    S.std1=std1; S.std2=std2
    S.taylor=array([sqrt(mean((x1-x1.mean())**2))/S.std2,sqrt(((x1-x1.mean()-x2+x2.mean())**2).mean())/S.std2,S.R])

    return S

#---------------------shpfile--------------------------------------------------
def read_shapefile_data(fname):
    '''
    read shapefile (*.shp) and return its content

    note:  works for pts and polygon only, may not work for other geomerties (need update in these cases)
    '''

    import shapefile as shp
    with shp.Reader(fname) as C:
        #----read shapefile----------------
        S=zdata();
        S.nrec=C.numRecords
        S.type=C.shapeTypeName

        #----read pts----------------------------------------------------------
        #works for pts and polygon, may not work for other geomerty (need update in this case)
        S.xy=[];
        for i in arange(S.nrec):
            xyi=array(C.shape(i).points);
            parti=array(C.shape(i).parts,dtype='int');
            #insert nan for delimiter
            #to get original index: ind=nonzero(isnan(xyi[:,0]))[0]-arange(len(parti));
            S.xy.append(insert(xyi,parti,nan,axis=0))
        S.xy=squeeze(array(S.xy))

        #---read attributes----------------------------------------------------
        S.attname=array([C.fields[i][0] for i in arange(1,len(C.fields))]);
        stype=array([type(C.record()[m]) for m in S.attname])
        svalue=array(C.records(),dtype='O');
        S.attvalue=array(zeros(len(S.attname))).astype('O')
        for m in arange(len(S.attname)):
            S.attvalue[m]=svalue[:,m].astype(stype[m])
        S.atttype=stype

        #read prj file if exist---
        bdir=os.path.dirname(os.path.abspath(fname));
        bname=os.path.basename(fname).split('.')[0]
        prjname='{}/{}.prj'.format(bdir,bname)
        if os.path.exists(prjname):
            with open(prjname,'r') as fid:
                S.prj=fid.readline().strip()

    return S

def write_shapefile_data(fname,S,float_len=18,float_decimal=8):
    '''
    write shapefile
        fname: file name
        S: data to be outputed

    example of S:
       S.type=='POINT'
       S.xy=c_[slon[:],slat[:]]
       S.prj=get_prj_file('epsg:4326')
       S.attname=['station']
       S.attvalue=station[:]

    note: only works for geometry: POINT, POLYLINE, POLYGON
    '''

    import shapefile as shp
    #---get nrec-----
    if S.type=='POINT':
        if S.xy.dtype==dtype('O'):
            print('S.xy has a dtype="O" for POINT'); sys.exit()
        else:
            nrec=S.xy.shape[0];
    elif S.type=='POLYLINE' or S.type=='POLYGON':
        if S.xy.dtype==dtype('O'):
            nrec=len(S.xy)
        else:
            nrec=1;
    else:
        sys.exit('unknow type')

    #---check nrec
    if hasattr(S,'nrec'):
        if nrec!=S.nrec:
            print('nrec inconsistent')
            sys.exit()

    #---write shapefile---------
    with shp.Writer(fname) as W:
        W.autoBalance=1;
        #define attributes
        if hasattr(S,'attname'):
            if S.attvalue.dtype==dtype('O'):
                stype=[type(S.attvalue[m][0]) for m in arange(len(S.attname))]
            else:
                if S.attvalue.ndim==1:
                    stype=[type(S.attvalue[0])]
                elif S.attvalue.ndim==2:
                    stype=[type(S.attvalue[m][0]) for m in arange(len(S.attname))]

            for m in arange(len(stype)):
                if stype[m] in [np.int,np.int8,np.int16,np.int32,np.int64]:
                    W.field(S.attname[m],'N')
                elif stype[m] in [np.float,np.float16,np.float32,np.float64]:
                    W.field(S.attname[m],'F',float_len,float_decimal)
                elif stype[m] in [np.str0,np.str,np.str_,np.string_]:
                    W.field(S.attname[m],'C',100)
                else:
                    print('attribute type not included: {}'.format(stype[m]))
                    sys.exit()
        else:
            W.field('field','C')
            W.record()

        #put values
        for i in arange(nrec):
            if S.type=='POINT': #point, W.multipoint(S.xy) is multiple pts features
                vali=S.xy[i]
                W.point(*vali)
            elif S.type=='POLYLINE':
                if S.xy.dtype==dtype('O'):
                    vali=S.xy[i]
                else:
                    vali=S.xy
                #reorganize the shape of vali
                valii=delete_shapefile_nan(vali,0)
                W.line(valii)
            elif S.type=='POLYGON':
                if S.xy.dtype==dtype('O'):
                    vali=S.xy[i]
                else:
                    vali=S.xy
                #reorganize the shape of vali
                valii=delete_shapefile_nan(vali,1)
                valii=[[[*k] for k in valii[0]]]
                W.poly(valii)

            #add attribute
            if hasattr(S,'attname'):
                if S.attvalue.dtype==dtype('O'):
                    atti=[S.attvalue[m][i] for m in arange(len(stype))]
                else:
                    if S.attvalue.ndim==1:
                        atti=[S.attvalue[i]]
                    elif S.attvalue.ndim==2:
                        atti=[S.attvalue[m][i] for m in arange(len(stype))]
                W.record(*atti)

        #----write projection------------
        bname=os.path.basename(fname).split('.')[0]
        bdir=os.path.dirname(os.path.abspath(fname));
        if hasattr(S,'prj'):
            with open('{}/{}.prj'.format(bdir,bname),'w+') as fid:
                fid.write(S.prj)

def delete_shapefile_nan(xi,iloop=0):
    '''
    delete nan (head and tail), and get ind for the rest
    '''
    if xi.ndim==1:
        i1=0; i2=xi.shape[0]
        if isnan(xi[0]): i1=1
        if isnan(xi[-1]): i2=i2-1
        yi=xi[i1:i2]; ind=nonzero(isnan(yi))[0]
    elif xi.ndim==2:
        i1=0; i2=xi.shape[0]
        if isnan(xi[0,0]): i1=1
        if isnan(xi[-1,0]): i2=i2-1
        yi=xi[i1:i2]; ind=nonzero(isnan(yi[:,0]))[0]

    #------reorganize-----------
    if len(ind)==0:
        #close the geomety
        if iloop==1: yi=close_data_loop(yi)

        vi=[yi];
    else:
        vi=[];
        yii=yi[:ind[0]];
        if iloop==1: yii=close_data_loop(yii)
        vi.append(yii)
        for m in arange(len(ind)-1):
            i1=ind[m]+1; i2=ind[m+1];
            yii=yi[i1:i2];
            if iloop==1: yii=close_data_loop(yii);
            vi.append(yii)
        yii=yi[(ind[-1]+1):];
        if iloop==1: yii=close_data_loop(yii)
        vi.append(yii)

    return vi

#---------------------netcdf---------------------------------------------------
def ReadNC(fname,fmt=0,order=0):
    '''
    read netcdf files, and return its values and attributes
        fname: file name

        fmt=0: reorgnaized Dataset with format of zdata
        fmt=1: return netcdf.Dateset(fname)
        fmt=2: reorgnaized Dataset (*npz format), ignore attributes

        order=1: only works for med=2; change dimension order
        order=0: variable dimension order read not changed for python format
        order=1: variable dimension order read reversed follwoing in matlab/fortran format
    '''

    #get file handle
    C=Dataset(fname);

    if fmt==1:
        return C
    elif fmt in [0,2]:
        F=zdata(); F.file_format=C.file_format; F.VINFO=[]

        #read dims
        ncdims=[i for i in C.dimensions];
        F.dimname=ncdims; F.dims=[]; F.dim_unlimited=[]
        for i in ncdims:
            F.dims.append(C.dimensions[i].size)
            F.dim_unlimited.append(C.dimensions[i].isunlimited())
        F.VINFO.append('dimname: {}'.format(F.dimname))
        F.VINFO.append('dim: {}'.format(F.dims))

        #read attrbutes
        ncattrs=C.ncattrs(); F.attrs=ncattrs
        for i in ncattrs:
            exec('F.{}=C.getncattr("{}")'.format(i,i))

        ncvars=[*C.variables]; F.vars=array(ncvars)
        #read variables
        for i in ncvars:
            if fmt==0:
               fi=zdata()
               dimi=C.variables[i].dimensions
               fi.dimname=dimi
               fi.dims=[C.dimensions[j].size for j in dimi]
               fi.val=C.variables[i][:]
               fi.attrs=C.variables[i].ncattrs()
               for j in C.variables[i].ncattrs():
                   ncattri=C.variables[i].getncattr(j)
                   exec('fi.{}=ncattri'.format(j))

               if order==1:
                   fi.dimname=list(flipud(fi.dimname))
                   fi.dims=list(flipud(fi.dims))
                   nm=flipud(arange(ndim(fi.val)))
                   fi.val=fi.val.transpose(nm)
               vinfo='{}:{}'.format(i,fi.val.shape)
            elif fmt==2:
               fi=array(C.variables[i][:])
               if order==1: fi=fi.transpose(flip(arange(fi.ndim)))
               vinfo='{}:{}'.format(i,fi.shape)
            F.VINFO.append(vinfo);  exec('F.{}=fi'.format(i));

        C.close(); #F.VINFO=array(F.VINFO)
        return F
    else:
        sys.exit('wrong fmt')

def WriteNC(fname,data,fmt=0,order=0):
    '''
    write zdata to netcdf file
        fname: file name
        data:  soure data
        fmt=0, data has zdata() format
        fmt=1, data has netcdf.Dataset format
        order=0: variable dimension order written not changed for python format
        order=1: variable dimension order written reversed follwoing in matlab/fortran format
    '''

    #pre-processing fname
    #if fname.endswith('.nc'): fname=fname[:-3]

    if fmt==1:
        #----write NC files-------------
        #fid=Dataset('{}.nc'.format(fname),'w',format=data.file_format); #C.file_format
        fid=Dataset(fname,'w',format=data.file_format); #C.file_format
        fid.setncattr('file_format',data.file_format)

        #set attrs
        ncattrs=data.ncattrs()
        for i in ncattrs:
            exec("fid.setncattr('{}',data.{})".format(i,i))

        #set dim
        ncdims=[i for i in data.dimensions]
        for i in ncdims:
            if data.dimensions[i].isunlimited() is True:
               fid.createDimension(i,None)
            else:
               fid.createDimension(i,data.dimensions[i].size)

        #set variable
        ncvars=[i for i in data.variables]
        if order==0:
            for vari in ncvars:
                vid=fid.createVariable(vari,data.variables[vari].dtype,data.variables[vari].dimensions)
                for attri in data.variables[vari].ncattrs():
                    vid.setncattr(attri,data.variables[vari].getncattr(attri))
                fid.variables[vari][:]=data.variables[vari][:]
        elif order==1:
            for vari in ncvars:
                vid=fid.createVariable(vari,data.variables[vari].dtype,flipud(data.variables[vari].dimensions))
                for attri in data.variables[vari].ncattrs():
                    vid.setncattr(attri,data.variables[vari].getncattr(attri))
                nm=flipud(arange(ndim(data.variables[vari][:])));
                fid.variables[vari][:]=data.variables[vari][:].transpose(nm)

        fid.close()
    elif fmt==0:
        #----write NC files-------------
        #fid=Dataset('{}.nc'.format(fname),'w',format=data.file_format); #C.file_format
        fid=Dataset(fname,'w',format=data.file_format); #C.file_format

        #set attrs
        fid.setncattr('file_format',data.file_format)
        if hasattr(data,'attrs'):
           for i in data.attrs:
               exec("fid.setncattr('{}',data.{})".format(i,i))

        #set dimension
        for i in range(len(data.dims)):
            if hasattr(data,'dim_unlimited'):
               dim_flag=data.dim_unlimited[i]
            else:
               dim_flag=False

            if dim_flag is True:
               fid.createDimension(data.dimname[i],None)
            else:
               fid.createDimension(data.dimname[i],data.dims[i])

        #set variable
        if order==0:
            for vari in data.vars:
                vi=eval('data.{}'.format(vari));
                vid=fid.createVariable(vari,vi.val.dtype,vi.dimname)
                if hasattr(vi,'attrs'):
                   for j in vi.attrs:
                       attri=eval('vi.{}'.format(j))
                       vid.setncattr(j,attri)
                fid.variables[vari][:]=vi.val
        elif order==1:
            for vari in data.vars:
                vi=eval('data.{}'.format(vari));
                vid=fid.createVariable(vari,vi.val.dtype,flipud(vi.dimname))
                if hasattr(vi,'attrs'):
                   for j in vi.attrs:
                       attri=eval('vi.{}'.format(j))
                       vid.setncattr(j,attri)
                if ndim(vi.val)>=2:
                    nm=flipud(arange(ndim(vi.val)));
                    fid.variables[vari][:]=vi.val.transpose(nm)
                else:
                    fid.variables[vari][:]=vi.val
        fid.close()

def harmonic_fit(oti,oyi,dt,mti,tidal_names=0, **args):
    '''
    construct time series using harmonics: used to fill data gap, or extrapolation
       [oti,oyi,dt]: observed (time, value, time interval)
       mti: time to interpolate/extrapolate
       tidal_names=0/1/2: tidal names options; tidal_names=['O1','K1','Q1','P1','M2','S2',]: list of tidal consituent names
    '''
    #choose tidal consts
    if not hasattr(tidal_names,'__len__'):
       if tidal_names==0:
          tidal_names=array(['O1','K1','Q1','P1','M2','S2','K2','N2'])
       elif tidal_names==1:
          tidal_names=array(['O1','K1','Q1','P1','M2','S2','K2','N2','M3','M4','M6','M7','M8','M10'])
       elif tidal_names==2:
          tidal_names=array(['SA','SSA','MM','MSF','O1','K1','Q1','P1','M2','S2','K2','N2','M3','M4','M6','M7','M8','M10','N4','S4','S6'])

    #do harmnoic analysis
    H=harmonic_analysis(oyi,dt,tidal_names=tidal_names,**args)

    #fit data
    myi=zeros(mti.shape)+H.amplitude[0]
    for k,tname in enumerate(tidal_names): myi=myi+H.amplitude[k+1]*cos(H.freq[k+1]*(mti-oti[0])*86400-H.phase[k+1])

    return myi

def harmonic_analysis(data,dt,t0=0,tidal_names=None,code=None,tname=None,fname=None,sname=None):
    '''
    harmonic analyze time series
        data: time series
        dt: time step (day)
        t0: starting time of time series (day); normally use datenum (e.g. datenum(2010,0,0))
        tidal_names: 1) path of tidal_const.dat, or 2) names of tidal consituents
        code: path of executable (tidal_analyze, or tidal_analyze.exe) for HA analysis
        [tname,fname,sname]: temporary names for tidal_const, time series, and HA results
    '''
    import subprocess

    #names of temporary file
    if tname is None: tname='.temporary_tidal_const_for_HA.dat'
    if fname is None: fname='.temporary_time_series_for_HA.dat'
    if sname is None: sname='.temporary_tidal_consituents_for_HA.dat'

    #check OS type, and locate the executable
    sdir='{}/../pyScripts/Harmonic_Analysis'.format(os.path.dirname(__file__))
    if code is None:
        import platform
        #directories where exectuable may exist
        bdirs=[sdir, r'D:\Work\Harmonic_Analysis',r'~/bin/Harmonic_Analysis']
        if platform.system().lower()=='windows':
           code=['{}/tidal_analyze.exe'.format(i) for i in bdirs if os.path.exists('{}/tidal_analyze.exe'.format(i))][0]
        elif platform.system().lower()=='linux':
           code=['{}/tidal_analyze'.format(i) for i in bdirs if os.path.exists('{}/tidal_analyze'.format(i))][0]
        else:
            print('Operating System unknow: {}'.format(platform.system()))
        if code is None: sys.exit('exectuable "tidal_analyze" was not found')

    #locate or write tidal_const.dat
    if tidal_names is None:
       tidal_names='{}/tidal_const.dat'.format(sdir)
    else:
       if not isinstance(tidal_names,str): #names of tidal consituents
          C=loadz('{}/tide_fac_const.npz'.format(sdir)); tdict=dict(zip(C.name,C.freq))
          fid=open(tname,'w+'); fid.write('{}\n'.format(len(tidal_names)))
          for i in tidal_names: fid.write('{}\n{}\n'.format(i.upper(),tdict[i.upper()]))
          fid.close(); tidal_names=tname

    #write time series, HA, and return results
    fid=open(fname,'w+')
    for i,datai in enumerate(data): fid.write('{:12.1f} {:12.7f}\n'.format((t0+i*dt)*86400,datai))
    fid.close(); subprocess.call([code,fname,tidal_names,sname]) #HA

    #save results
    S=zdata(); S.tidal_name,S.amplitude,S.phase=array([i.strip().split() for i in open(sname,'r').readlines()]).T
    S.amplitude=array(S.amplitude).astype('float'); S.phase=array(S.phase).astype('float')
    S.freq=r_[0,array([i.strip() for i in open(tidal_names,'r').readlines()[2::2]]).astype('float')]

    #clean temporaray files--------
    os.remove(fname); os.remove(sname)
    if os.path.exists(tname): os.remove(tname)
    return S

def get_hycom(Time,xyz,vind,hdir='./HYCOM',method=0):
    '''
    extract Hycom time series at stations
    ti: time seires; xyz=c_[loni,lati,depi];
    vind: list of index for variables to be extracted. [0,1,2,3] for ['elev','temp','salt','uv']
    hdir: directory for hycom data
    method=0: linear interpolation; method=1: nearest interpolation
    '''

    #variable names
    Var=['surf_el','water_temp','salinity',['water_u','water_v']];
    VarName=['elev','temp','salt',['Ux','Uy']];

    #time
    StartT=floor(Time.min())
    EndT=ceil(Time.max())

    #---interpolation pts----
    if xyz.ndim==1: xyz=xyz[None,:]
    loni=xyz[:,0]; lati=xyz[:,1]
    bxy=c_[lati,loni];
    if xyz.shape[1]==3:
       depi=xyz[:,2]
       bxyz=c_[depi,lati,loni];

    #-----save data----------------
    S=zdata();
    S.x=xyz[:,0]; S.y=xyz[:,1]
    if xyz.shape[1]==3: S.z=xyz[:,2]

    #---read Hycom data and interpolate onto boundary nodes------------------
    p=zdata();
    for i in vind: #arange(len(Var)):
        vari=Var[i]; varnamei=VarName[i];

        t0=time.time();
        T0=[]; Data0=[];
        for ti in arange(StartT,EndT+1):
            t1=num2date(ti); t2=num2date(ti+1-1/24/60);

            if isinstance(vari,list):
                fname='Hycom_{}_{}_{}.nc'.format(varnamei[0],t1.strftime('%Y%m%dZ%H%M00'),t2.strftime('%Y%m%dZ%H%M00'))
                fname2='Hycom_{}_{}_{}.nc'.format(varnamei[1],t1.strftime('%Y%m%dZ%H%M00'),t2.strftime('%Y%m%dZ%H%M00'))
                if not os.path.exists(r'{}/{}'.format(hdir,fname)): continue
                if not os.path.exists(r'{}/{}'.format(hdir,fname2)): continue
                print(fname+'; '+fname2)
                C=ReadNC('{}/{}'.format(hdir,fname)); C2=ReadNC('{}/{}'.format(hdir,fname2),2)

                #get value
                exec('p.val=C.{}.val'.format(vari[0]))
                exec('p.val2=C2.{}.val'.format(vari[1]))
                p.val=array(p.val); fp=p.val<-29999; p.val2=array(p.val2); fp2=p.val2<-29999;
                p.val[fp]=nan; p.val2[fp2]=nan;
            else:
                fname='Hycom_{}_{}_{}.nc'.format(varnamei,t1.strftime('%Y%m%dZ%H%M00'),t2.strftime('%Y%m%dZ%H%M00'))
                if not os.path.exists(r'{}/{}'.format(hdir,fname)): continue
                print(fname)
                C=ReadNC('{}/{}'.format(hdir,fname))

                #get value
                exec('p.val=C.{}.val'.format(vari))
                p.val=array(p.val); fp=p.val<-29999;
                p.val[fp]=nan

            ti=datestr2num(C.time.time_origin)+array(C.time.val)/24
            cloni=array(C.lon.val); cloni=mod(cloni,360)-360
            clati=array(C.lat.val);

            #------define data region extracted
            ind_lon=nonzero((cloni<=max(loni)+0.1)*(cloni>=min(loni)-0.1))[0];
            ind_lat=nonzero((clati<=max(lati)+0.1)*(clati>=min(lati)-0.1))[0];
            i1_lon=ind_lon.min(); i2_lon=i1_lon+len(ind_lon)
            i1_lat=ind_lat.min(); i2_lat=i1_lat+len(ind_lat)

            cloni=cloni[i1_lon:i2_lon]; clati=clati[i1_lat:i2_lat]

            if varnamei=='elev':
                for m in arange(len(ti)):
                    valii=squeeze(p.val[m,i1_lat:i2_lat,i1_lon:i2_lon])

                    if method==0:
                        #interpolation
                        fd=sp.interpolate.RegularGridInterpolator((clati,cloni),valii,fill_value=nan)
                        vi=fd(bxy)

                        #remove nan pts
                        fp=isnan(vi);
                        if sum(fp)!=0:
                            vi[fp]=sp.interpolate.griddata(bxy[~fp,:],vi[~fp],bxy[fp,:],'nearest')

                    elif method==1:
                        clonii,clatii=meshgrid(cloni,clati); sind=~isnan(valii.ravel())
                        clonii=clonii.ravel()[sind]; clatii=clatii.ravel()[sind]; valii=valii.ravel()[sind]
                        vi=sp.interpolate.griddata(c_[clatii,clonii],valii,bxy,'nearest')
                        # scatter(clonii,clatii,s=8,c=valii); plot(bxy[:,1],bxy[:,0],'r.'); sys.exit()
                    T0.append(ti[m]); Data0.append(vi);
            else:
                #------define data region extracted for depth
                cdepi=array(C.depth.val)
                ind_dep=nonzero((cdepi<=depi.max()+1000)*(cdepi>=depi.min()-100))[0];
                i1_dep=ind_dep.min(); i2_dep=i1_dep+len(ind_dep)
                cdepi=cdepi[i1_dep:i2_dep];

                for m in arange(len(ti)):
                    T0.append(ti[m]);
                    valii=squeeze(p.val[m,i1_dep:i2_dep,i1_lat:i2_lat,i1_lon:i2_lon])

                    #interpolation
                    fd=sp.interpolate.RegularGridInterpolator((cdepi,clati,cloni),valii,fill_value=nan)
                    vi=fd(bxyz)

                    #remove nan pts
                    fp=isnan(vi);
                    if sum(fp)!=0:
                        vi[fp]=sp.interpolate.griddata(bxyz[~fp,:],vi[~fp],bxyz[fp,:],'nearest')

                    #----if variable is velocity
                    if isinstance(varnamei,list):
                        val2ii=squeeze(p.val2[m,i1_dep:i2_dep,i1_lat:i2_lat,i1_lon:i2_lon])

                        #interpolation
                        fd=sp.interpolate.RegularGridInterpolator((cdepi,clati,cloni),val2ii,fill_value=nan)
                        v2i=fd(bxyz)

                        #remove nan pts
                        fp=isnan(v2i);
                        if sum(fp)!=0:
                            v2i[fp]=sp.interpolate.griddata(bxyz[~fp,:],v2i[~fp],bxyz[fp,:],'nearest')

                        Data0.append(r_[expand_dims(vi,0),expand_dims(v2i,0)])

                    else:
                        Data0.append(vi)

        #interpolate in time, and save data
        T0=array(T0); Data0=array(Data0).T; S.time=Time;
        if Data0.ndim==2:
            #interpolate
            datai=[];
            for k in arange(Data0.shape[0]):
                fd=interpolate.interp1d(T0,Data0[k],fill_value='extrapolate');
                datai.append(fd(Time));
            datai=array(datai)
            exec('S.{}=datai'.format(varnamei))
        else:
            #interpolate
            datai=[]; data2i=[];
            for k in arange(Data0.shape[0]):
                fd=interpolate.interp1d(T0,Data0[k,0]);
                fd2=interpolate.interp1d(T0,Data0[k,1]);
                datai.append(fd(Time)); data2i.append(fd2(Time))
            datai=array(datai); data2i=array(data2i)
            exec('S.{}=datai'.format(varnamei[0]))
            exec('S.{}=data2i'.format(varnamei[1]))

    return S



if __name__=="__main__":
    pass

##---------get_hycom------------------------------------------------------------
#    #----inputs--------------
#    StartT=datenum(2014,1,1)
#    EndT=datenum(2014,1,4)
#    ti=arange(StartT,EndT,1/24);
#
#    #--convert coordinate if not lon&lat---
#    bp=read_schism_bpfile('Station.bp_WQ')
#    loni,lati=transform(Proj(init='epsg:26919'),Proj(init='epsg:4326'),bp.x,bp.y);
#    xyz=c_[loni,lati,zeros(loni.shape)]
#    xyz=array([-65.91084725,  42.33315317,   0.])
#
#    vind=[0,1,2,3]; #variable index for extract, vind=['elev','temp','salt','uv']
#
#    S=get_hycom(ti,xyz,vind,hdir='./Data');

##-------------HA --------------------------------------------------------------
#    data=loadtxt(r'D:\Work\Harmonic_Analysis\TS.txt');
#    S=harmonic_analysis(data[:,1],1/24,tidal_const=r'D:\Work\Harmonic_Analysis\tmp.dat')
#    S=harmonic_analysis(data[:,1],1/24)

#---------------------projection-----------------------------------------------
#    fname0=r'D:\Work\E3SM\ChesBay\Grid\ChesBay_1a.gr3';
#    fname1=r'D:\Work\E3SM\ChesBay\Grid\ChesBay_1a.ll';
#    proj(fname0,0,'epsg:26918',fname1,3,'epsg:26918')
#    proj(fname0,0,'epsg:26918',fname1,0,'epsg:4326')
#    proj(fname0,3,'epsg:26918',fname1,2,'epsg:4326')

#------------------------------------------------------------------------------
#    clear_globals()
#------------------------------------------------------------------------------
#    xi=arange(100)*0.1;
#    yi=np.random.rand(100)+sin(xi)
#    y1=smooth(yi,3); y2=smooth(yi,5); y3=smooth(yi,7)
#    plot(xi,yi,'r-',xi,y1,'b-',xi,y2,'g-',xi,y3,'k-')
#    show()

#------------------str2num-----------------------------------------------------
#        x='3.5, 4, 5; 5  6.5, 78';
#    x='3.5 4 5 5 6.5   78'
#        xi=str2num(x);
#    x=['3 4 5','4 3 6 8']
#    x=['3 4 5','4 3 6']
#    xi=str2num(x)
#    print(xi)

#     #--test files--
#     fname=r'C:\Users\Zhengui\Desktop\Python\learn\hgrid.gr3'
#     with open(fname,'r') as fid:
#         lines=fid.readlines()
#     line=lines[24535:24545]
#     rline=remove_tail(line)
#     print(line)
#     print(rline)

#-------date_proc---------------------------------------------------------------
#    n1=(2006.,1,1)
#    n2=array([2006,2,1]);
#    n3=array([[2006,3,2,1,1],[2006,3,3,3,0]]);
#
#    f1=datenum(*n1);
#    f2=datenum(n2);
#    f3=datenum(n3,doy=1);

#-------mfft--------------------------------------------------------------------
#    plt.close('all')
#    T=10; dt=0.01; N=T/dt;
#    x=linspace(0.0, T, N);
#    y=4*cos(2.0*pi*(x-0.3)/0.5)+2*cos(2.0*pi*(x-0.4)/1.0)+4*cos(2.0*pi*(x-0.5)/2.0)
#    f,a,p=mfft(y,dt)
#
#    subplot(2,1,1)
#    plot(x,y,'k-')
#    subplot(2,1,2)
#    plot(f,a,'k.',ms=20)
#    setp(gca(),'xlim',[0,5])

#---------------------shpfile--------------------------------------------------

#---------------------netcdf---------------------------------------------------
#    # read NC files
#    C=Dataset('sflux_air_1.002.nc')
#
#    ncdims=[i for i in C.dimensions]
#    ncvars=[i for i in C.variables]
#
#
#    [print("{}".format(i)) for i in ncdims]
#    [print("{}".format(i)) for i in ncvars]
#
#    #----write NC files-------------
#    fid=Dataset('test.nc','w',format='NETCDF3_CLASSIC'); #C.file_format
#
#    for dimi in ncdims:
#        fid.createDimension(dimi,C.dimensions[dimi].size)
#
#    for vari in ncvars:
#        vid=fid.createVariable(vari,C.variables[vari].dtype,C.variables[vari].dimensions)
#        for attri in C.variables[vari].ncattrs():
#           vid.setncattr(attri,C.variables[vari].getncattr(attri))
#        fid.variables[vari][:]=C.variables[vari][:]
#    fid.close()
#
#    ## check results
#    F=Dataset('test.nc');

#----------------convert_matfile_format----------------------------------------
#    fname=r'C:\Users\Zhengui\Desktop\Observation2\USGS\SFBay_USGSData_MAL.mat'
#    fname=r'C:\Users\Zhengui\Desktop\convert_matfile\tem.mat'
#    cmat.convert_matfile_format(fname)


#------------------------------------------------------------------------------
##################code outdated: just for reference############################
#------------------------------------------------------------------------------
# def lpfilt(data,delta_t,cutoff_f):
#     '''
#     low pass filter for 1D (data[time]) or 2D (data[time,array]) array;
#     '''
#     #import gc #discard


#     ds=data.shape
#     mn=data.mean(axis=0)
#     data=data-mn
#     P=fft(data,axis=0)

#     #desgin filter
#     N=ds[0];
#     filt=ones(N)
#     k=int(floor(cutoff_f*N*delta_t))
#     filt[k]=0.715
#     filt[k+1]=0.24
#     filt[k+2]=0.024
#     filt[k+3:N-(k+4)]=0.0
#     filt[N-(k+4)]=0.024
#     filt[N-(k+3)]=0.24
#     filt[N-(k+2)]=0.715

#     #expand filter dimensions
#     fstr=',None'*(len(ds)-1); s=zdata()
#     exec('s.filt=filt[:{}]'.format(fstr))

#     #remove high freqs
#     P=P*s.filt

#     #lp results
#     fdata=real(ifft(P,axis=0))+mn

#     return fdata

# def wipe(n=50):
#     print('\n'*n)
#     return

#def clear_globals():
#    allvar = [var for var in globals() if var[0] != "_"]
#    for var in allvar:
#       #global var
#       #del var;
#       #del globals()[var]
#       #exec('del global()['+var+']')
#       exec('global '+var)
#       exec('del '+var)

#def reload(gfunc):
#    #reload modules,mainly used for coding debug
#    #usage: reload(globals())
#    import inspect,imp
#    mods=['mylib','schism_file','mpas_file']
#    for modi in mods:
#        imp.reload(sys.modules[modi])
#        #get all module functions
#        fs=[];afs=inspect.getmembers(sys.modules[modi],inspect.isfunction);
#        for fsi in afs:
#            if inspect.getmodule(fsi[1]).__name__!=modi: continue
#            if fsi[0] not in gfunc.keys(): continue
#            fs.append(fsi)
#        #refresh module functions
#        for fsi in fs:
#            if gfunc[fsi[0]]!=fsi[1]: gfunc[fsi[0]]=fsi[1]
#    return

#def near_pts(pts,pts0):
#    #pts[n,2]: xy of points
#    #pts0[n,2]: xy of points
#    #return index of pts0(x0,y0) that pts(x,y) is nearest
#    x=pts[:,0]; y=pts[:,1]
#    x0=pts0[:,0]; y0=pts0[:,1]
#    dist=(x[None,:]-x0[:,None])**2+(y[None,:]-y0[:,None])**2
#
#    ind=[];
#    for i in arange(x.shape[0]):
#        disti=dist[:,i];
#        ind.append(nonzero(disti==min(disti))[0][0])
#    ind=array(ind);
#
#    return ind

#def compute_cofficient(myi,oyi):
#    #compute different evaluation coefficients
#    N=len(myi)
#    mmyi=myi.mean(); moyi=oyi.mean();
#    emyi=myi-mmyi; eoyi=oyi-moyi; e=myi-oyi
#    stdm=std(emyi); stdo=std(eoyi)
#
#    SS_tot=sum((oyi-moyi)**2)
#    SS_reg=sum((myi-moyi)**2);
#    SS_res=sum(e**2);
#
#    #evaluation coefficient
#    ms=1-SS_res/sum((abs(myi-moyi)+abs(oyi-moyi))**2) #model skill
#    r=mean((myi-mmyi)*(oyi-moyi))/stdm/stdo #correlation coefficient
#    r2=1-SS_res/SS_tot  #R2
#    rmse=sqrt(sum(e**2)/N) #RMSE
#    mae=mean(abs(e))         #MAE
#    me=mean(e)               #ME
#
#    return [ms,r,r2,rmse,mae,me]

#-------str2num----------------------------------------------------------------
#def str2num(line,*args):
#    num=str2num_process(line,*args)
#    if isinstance(num[0],float):
#        num=num.astype('float64')
#    else:
#        num=[s.astype('float64') for s in num]
#        num=array(num)
#    return num
#
#@np.vectorize
#def str2num_process(line,*args):
#    if len(args)>0:
#        if len(args)>1:
#            for i in range(len(args)-1):
#                line=line.replace(arg)
#        line=line.replace(args[0],',')
#    else:
#        line=line.replace(';',',').replace(' ',',')
#    linei=[s for s in line.split(',') if s]
#    fc=np.vectorize(lambda x: np.float64(x))
#    return fc(linei).astype('object')
#
#@np.vectorize
#def remove_tail(line):
#    li=line.rstrip();
#    ind=li.find('!');
#    if ind!=-1:
#        li=li[:ind]
#    ind=li.find('=');
#    if ind!=-1:
#        li=li[:ind]
#    return li

#-------date_proc--------------------------------------------------------------
#def datenum_0(*args):
#    if len(args)==1:
#        args=args[0];
#
#    args=array(args)
#    args=args.astype('int')
#    return datetime.datetime(*args)
#
#def datenum(*args,doy=0):
#    #usage: datenum(2001,1,1)
#    args=array(args)
#    e1=args[0]
#
#    if hasattr(e1, "__len__"):
#        if not hasattr(e1[0],"__len__"):
#            f=datenum_0(*e1)
#        else:
#            f=apply_along_axis(datenum_0,1,e1)
#    else:
#        f=datenum_0(*args)
#    if doy==0:
#        return date2num(f)
#    else:
#        return f
#------------------------------------------------------------------------------

#-------smooth-----------------------------------------------------------------
# def smooth(xi,N):
#     '''
#     smooth average:
#        xi: time series
#        N: window size (if N is even, then N=N+1)
#     '''

#     if mod(N,2)==0: N=N+1
#     nz=int((N-1)/2)

#     X=xi; SN=ones(len(xi))*N
#     for i in arange(nz):
#         nmove=i+1; ii=-i-1;
#         xext=zeros(nmove)
#         X=X+r_[xext,xi[:ii]]
#         for j in range(nmove):
#             SN[j]=SN[j]-1

#     for i in arange(nz):
#         nmove=i+1; ii=i+1
#         xext=zeros(nmove)
#         X=X+r_[xi[ii:],xext]
#         for j in arange(nmove):
#             jj=-j-1
#             SN[jj]=SN[jj]-1
#     SX=np.divide(X,SN)
#     return SX
#------------------------------------------------------------------------------
